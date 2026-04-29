"""
MoneyMap – Reports Blueprint
Generates downloadable PDF and Excel reports from Firestore data.
"""
from flask import Blueprint, request, session, redirect, url_for, send_file
from functools import wraps
from datetime import datetime, date
import io
import calendar

# PDF
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

# Excel
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

from firebase_db import query_docs
from config import *

reports_bp = Blueprint('reports', __name__)

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated

def uid():
    return session['user_id']

def current_month():
    return date.today().strftime('%Y-%m')

def fmt_inr(amount):
    return f"₹{amount:,.2f}"

# ─────────────────────────────────────────────────────────────────────
# Helper – gather report data
# ─────────────────────────────────────────────────────────────────────

def _collect_data(month: str) -> dict:
    txns  = query_docs(COLL_TRANSACTIONS,  filters=[("user_id", "==", uid())])
    cats  = query_docs(COLL_CATEGORIES,    filters=[("user_id", "==", uid())])
    budg  = query_docs(COLL_BUDGETS,       filters=[("user_id", "==", uid()), ("month","==",month)])
    subs  = query_docs(COLL_SUBSCRIPTIONS, filters=[("user_id", "==", uid())])
    bills = query_docs(COLL_BILLS,         filters=[("user_id", "==", uid())])
    loans = query_docs(COLL_LOANS,         filters=[("user_id", "==", uid())])
    invs  = query_docs(COLL_INVESTMENTS,   filters=[("user_id", "==", uid())])

    cat_map = {c['id']: c for c in cats}

    month_txns = [t for t in txns if t.get('date', '').startswith(month)]
    income  = sum(t['amount'] for t in month_txns if t['type'] == 'income')
    expense = sum(t['amount'] for t in month_txns if t['type'] == 'expense')

    for t in month_txns:
        cat_id = t.get('category_id', '')
        t['category_name'] = cat_map.get(cat_id, {}).get('name', 'Uncategorised')

    # budget with spend
    for b in budg:
        spent = sum(
            t['amount'] for t in month_txns
            if t['type'] == 'expense' and t.get('category_id') == b.get('category_id')
        )
        b['spent'] = spent
        b['category_name'] = cat_map.get(b.get('category_id', ''), {}).get('name', 'N/A')

    # investment gain/loss
    for inv in invs:
        inv['gain'] = inv.get('current_val', 0) - inv.get('amount', 0)

    return {
        'month':       month,
        'income':      income,
        'expense':     expense,
        'balance':     income - expense,
        'transactions': sorted(month_txns, key=lambda x: x.get('date',''), reverse=True),
        'budgets':     budg,
        'subscriptions': subs,
        'bills':       bills,
        'loans':       loans,
        'investments': invs,
        'user_name':   session.get('user_name', 'User'),
    }

# ═══════════════════════════════════════════════════════════════════════
# PDF REPORT
# ═══════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/pdf')
@login_required
def download_pdf():
    month = request.args.get('month', current_month())
    data  = _collect_data(month)
    buf   = _build_pdf(data)
    fname = f"MoneyMap_Report_{month}.pdf"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')

def _build_pdf(data: dict) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm,  bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    PRIMARY = colors.HexColor('#4f46e5')
    SUCCESS = colors.HexColor('#10b981')
    DANGER  = colors.HexColor('#ef4444')
    LIGHT   = colors.HexColor('#f1f5f9')
    MUTED   = colors.HexColor('#64748b')

    title_style = ParagraphStyle('Title2', parent=styles['Title'],
                                 textColor=PRIMARY, fontSize=22, spaceAfter=6)
    h2_style    = ParagraphStyle('H2', parent=styles['Heading2'],
                                 textColor=PRIMARY, fontSize=13, spaceBefore=14, spaceAfter=6)
    normal      = styles['Normal']
    right_style = ParagraphStyle('Right', parent=normal, alignment=TA_RIGHT)

    story = []

    # ── Header ────────────────────────────────────────────────────────
    story.append(Paragraph("💰 MoneyMap", title_style))
    story.append(Paragraph(
        f"Financial Report – {data['month']} | Generated for {data['user_name']}",
        ParagraphStyle('Sub', parent=normal, textColor=MUTED, fontSize=10, spaceAfter=4)
    ))
    story.append(Paragraph(
        f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
        ParagraphStyle('Sub2', parent=normal, textColor=MUTED, fontSize=9, spaceAfter=8)
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=PRIMARY))
    story.append(Spacer(1, 12))

    # ── Summary cards ─────────────────────────────────────────────────
    story.append(Paragraph("Monthly Summary", h2_style))
    summary_data = [
        ["Income", "Expenses", "Net Balance"],
        [fmt_inr(data['income']),
         fmt_inr(data['expense']),
         fmt_inr(data['balance'])],
    ]
    tbl = Table(summary_data, colWidths=[5.5*cm, 5.5*cm, 5.5*cm])
    tbl.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), PRIMARY),
        ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 10),
        ('BACKGROUND',  (0,1), (0,1),  SUCCESS),
        ('BACKGROUND',  (1,1), (1,1),  DANGER),
        ('BACKGROUND',  (2,1), (2,1),  PRIMARY if data['balance'] >= 0 else DANGER),
        ('TEXTCOLOR',   (0,1), (-1,1), colors.white),
        ('FONTNAME',    (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,1), (-1,1), 13),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [None]),
        ('ROWHEIGHT',   (0,0), (-1,-1), 28),
        ('ROUNDEDCORNERS', [4]),
        ('BOX',         (0,0), (-1,-1), 0.5, colors.white),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 16))

    # ── Transactions ──────────────────────────────────────────────────
    story.append(Paragraph("Transactions", h2_style))
    if data['transactions']:
        rows = [["Date", "Category", "Note", "Type", "Amount"]]
        for t in data['transactions'][:30]:  # cap at 30
            amt_color = SUCCESS if t['type'] == 'income' else DANGER
            rows.append([
                t.get('date',''),
                t.get('category_name',''),
                t.get('note','')[:30],
                t['type'].upper(),
                fmt_inr(t['amount']),
            ])
        txn_tbl = Table(rows, colWidths=[2.5*cm, 3.5*cm, 5*cm, 2.2*cm, 3.3*cm])
        txn_style = [
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT]),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
            ('ALIGN',      (-1,0), (-1,-1), 'RIGHT'),
        ]
        # colour income/expense rows
        for i, t in enumerate(data['transactions'][:30], start=1):
            c = SUCCESS if t['type'] == 'income' else DANGER
            txn_style.append(('TEXTCOLOR', (4,i), (4,i), c))
            txn_style.append(('FONTNAME',  (4,i), (4,i), 'Helvetica-Bold'))
        txn_tbl.setStyle(TableStyle(txn_style))
        story.append(txn_tbl)
    else:
        story.append(Paragraph("No transactions this month.", normal))
    story.append(Spacer(1, 16))

    # ── Budgets ───────────────────────────────────────────────────────
    if data['budgets']:
        story.append(Paragraph("Budget vs Actual", h2_style))
        brows = [["Category", "Budget", "Spent", "Remaining", "Status"]]
        for b in data['budgets']:
            rem  = b['amount'] - b['spent']
            stat = "✅ OK" if rem >= 0 else "⚠️ Over"
            brows.append([
                b['category_name'],
                fmt_inr(b['amount']),
                fmt_inr(b['spent']),
                fmt_inr(rem),
                stat,
            ])
        btbl = Table(brows, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3.5*cm])
        btbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT]),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
            ('ALIGN',      (1,0), (-1,-1), 'RIGHT'),
        ]))
        story.append(btbl)
        story.append(Spacer(1, 16))

    # ── Subscriptions ────────────────────────────────────────────────
    if data['subscriptions']:
        story.append(Paragraph("Subscriptions", h2_style))
        srows = [["Name", "Amount/Month", "Renewal Day"]]
        total_sub = 0
        for s in data['subscriptions']:
            srows.append([s['name'], fmt_inr(s['amount']), f"Day {s.get('renewal_day',1)}"])
            total_sub += s['amount']
        srows.append(["TOTAL", fmt_inr(total_sub), ""])
        stbl = Table(srows, colWidths=[6*cm, 4*cm, 6.5*cm])
        stbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,-1), 'Helvetica'),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
            ('BACKGROUND', (0,-1), (-1,-1), LIGHT),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, LIGHT]),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ]))
        story.append(stbl)
        story.append(Spacer(1, 16))

    # ── Investments ───────────────────────────────────────────────────
    if data['investments']:
        story.append(Paragraph("Investments", h2_style))
        irows = [["Name", "Type", "Invested", "Current Value", "Gain/Loss"]]
        for inv in data['investments']:
            gain = inv.get('gain', 0)
            irows.append([
                inv['name'], inv['type'],
                fmt_inr(inv['amount']),
                fmt_inr(inv.get('current_val', 0)),
                fmt_inr(gain),
            ])
        itbl = Table(irows, colWidths=[4*cm, 3.5*cm, 3*cm, 3.5*cm, 2.5*cm])
        istyle = [
            ('BACKGROUND', (0,0), (-1,0), PRIMARY),
            ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT]),
            ('GRID',       (0,0), (-1,-1), 0.3, colors.HexColor('#e2e8f0')),
        ]
        for i, inv in enumerate(data['investments'], start=1):
            c = SUCCESS if inv.get('gain', 0) >= 0 else DANGER
            istyle.append(('TEXTCOLOR', (4,i), (4,i), c))
        itbl.setStyle(TableStyle(istyle))
        story.append(itbl)

    # ── Footer ────────────────────────────────────────────────────────
    story.append(Spacer(1, 24))
    story.append(HRFlowable(width="100%", thickness=1, color=MUTED))
    story.append(Paragraph(
        "Generated by MoneyMap • Personal Finance Management",
        ParagraphStyle('Footer', parent=normal, textColor=MUTED,
                       fontSize=8, alignment=TA_CENTER, spaceBefore=6)
    ))

    doc.build(story)
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════════════════════
# EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════

@reports_bp.route('/reports/excel')
@login_required
def download_excel():
    month = request.args.get('month', current_month())
    data  = _collect_data(month)
    buf   = _build_excel(data)
    fname = f"MoneyMap_Report_{month}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

def _build_excel(data: dict) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # ── Colour palette ────────────────────────────────────────────────
    C_PRIMARY = "4F46E5"
    C_SUCCESS = "10B981"
    C_DANGER  = "EF4444"
    C_LIGHT   = "F1F5F9"
    C_WHITE   = "FFFFFF"
    C_HEADER_TXT = "FFFFFF"
    C_MUTED   = "64748B"

    def header_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style='thin', color="E2E8F0")
        return Border(left=s, right=s, top=s, bottom=s)

    def header_font(bold=True):
        return Font(name='Calibri', bold=bold, color=C_HEADER_TXT, size=10)

    def normal_font(bold=False, color="1E293B"):
        return Font(name='Calibri', bold=bold, color=color, size=9)

    def center():
        return Alignment(horizontal='center', vertical='center', wrap_text=True)

    def right():
        return Alignment(horizontal='right', vertical='center')

    # ── Sheet helper ──────────────────────────────────────────────────
    def style_header_row(ws, row_num, col_count, fill_color=C_PRIMARY):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill        = header_fill(fill_color)
            cell.font        = header_font()
            cell.alignment   = center()
            cell.border      = thin_border()

    def style_data_row(ws, row_num, col_count, alt=False):
        fill = header_fill("F8FAFC") if alt else header_fill(C_WHITE)
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill      = fill
            cell.font      = normal_font()
            cell.border    = thin_border()
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # ─────────────────────────────────────────────────────────────────
    # Sheet 1 – Summary
    # ─────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value     = f"MoneyMap – Financial Report  |  {data['month']}"
    t.font      = Font(name='Calibri', bold=True, size=16, color=C_PRIMARY)
    t.alignment = center()
    t.fill      = header_fill("EEF2FF")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value     = f"User: {data['user_name']}  |  Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}"
    s.font      = Font(name='Calibri', size=9, color=C_MUTED)
    s.alignment = center()
    ws.row_dimensions[2].height = 20

    # Summary boxes
    for c, (label, value, clr) in enumerate([
        ("Total Income",  fmt_inr(data['income']),  C_SUCCESS),
        ("Total Expense", fmt_inr(data['expense']), C_DANGER),
        ("Net Balance",   fmt_inr(data['balance']), C_PRIMARY),
    ], start=1):
        col_start = (c - 1) * 2 + 1
        ws.merge_cells(start_row=4, start_column=col_start,
                       end_row=4,   end_column=col_start + 1)
        ws.merge_cells(start_row=5, start_column=col_start,
                       end_row=5,   end_column=col_start + 1)
        lbl_cell = ws.cell(row=4, column=col_start, value=label)
        val_cell = ws.cell(row=5, column=col_start, value=value)
        lbl_cell.fill = header_fill(clr)
        val_cell.fill = header_fill(clr)
        for cell in [lbl_cell, val_cell]:
            cell.font      = Font(name='Calibri', bold=True, color=C_WHITE,
                                   size=11 if cell == val_cell else 9)
            cell.alignment = center()
            cell.border    = thin_border()

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 28

    # Column widths
    for col, w in zip("ABCDE", [18, 18, 18, 18, 18]):
        ws.column_dimensions[get_column_letter(col if isinstance(col, int)
                                               else ord(col) - 64)].width = w

    # ─────────────────────────────────────────────────────────────────
    # Sheet 2 – Transactions
    # ─────────────────────────────────────────────────────────────────
    wt = wb.create_sheet("Transactions")
    wt.sheet_view.showGridLines = False
    headers = ["Date", "Category", "Note", "Type", "Amount (₹)"]
    for ci, h in enumerate(headers, 1):
        wt.cell(row=1, column=ci, value=h)
    style_header_row(wt, 1, len(headers))
    wt.row_dimensions[1].height = 22

    for ri, t in enumerate(data['transactions'], 2):
        alt = ri % 2 == 0
        row = [
            t.get('date', ''),
            t.get('category_name', ''),
            t.get('note', ''),
            t['type'].upper(),
            t['amount'],
        ]
        for ci, val in enumerate(row, 1):
            wt.cell(row=ri, column=ci, value=val)
        style_data_row(wt, ri, len(headers), alt)

        # Colour amount cell
        amt_cell  = wt.cell(row=ri, column=5)
        type_cell = wt.cell(row=ri, column=4)
        c = C_SUCCESS if t['type'] == 'income' else C_DANGER
        amt_cell.font  = normal_font(bold=True, color=c)
        amt_cell.number_format = '₹#,##0.00'
        amt_cell.alignment = right()

    for col, w in zip([1,2,3,4,5], [12, 16, 28, 10, 14]):
        wt.column_dimensions[get_column_letter(col)].width = w

    # ─────────────────────────────────────────────────────────────────
    # Sheet 3 – Budgets
    # ─────────────────────────────────────────────────────────────────
    wb2 = wb.create_sheet("Budgets")
    wb2.sheet_view.showGridLines = False
    bheaders = ["Category", "Budget (₹)", "Spent (₹)", "Remaining (₹)", "Status"]
    for ci, h in enumerate(bheaders, 1):
        wb2.cell(row=1, column=ci, value=h)
    style_header_row(wb2, 1, len(bheaders))

    for ri, b in enumerate(data['budgets'], 2):
        rem  = b['amount'] - b['spent']
        stat = "✅ Within Budget" if rem >= 0 else "⚠️ Over Budget"
        row  = [b['category_name'], b['amount'], b['spent'], rem, stat]
        for ci, val in enumerate(row, 1):
            wb2.cell(row=ri, column=ci, value=val)
        style_data_row(wb2, ri, len(bheaders), ri % 2 == 0)
        for col in [2, 3, 4]:
            wb2.cell(row=ri, column=col).number_format = '₹#,##0.00'
        rem_cell = wb2.cell(row=ri, column=4)
        rem_cell.font = normal_font(bold=True, color=C_SUCCESS if rem >= 0 else C_DANGER)

    for col, w in zip([1,2,3,4,5], [18, 14, 14, 14, 18]):
        wb2.column_dimensions[get_column_letter(col)].width = w

    # ─────────────────────────────────────────────────────────────────
    # Sheet 4 – Subscriptions
    # ─────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Subscriptions")
    ws2.sheet_view.showGridLines = False
    sheaders = ["Subscription", "Monthly Amount (₹)", "Renewal Day"]
    for ci, h in enumerate(sheaders, 1):
        ws2.cell(row=1, column=ci, value=h)
    style_header_row(ws2, 1, len(sheaders))

    total_sub = 0
    for ri, s in enumerate(data['subscriptions'], 2):
        row = [s['name'], s['amount'], f"Day {s.get('renewal_day',1)}"]
        for ci, val in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=val)
        style_data_row(ws2, ri, len(sheaders), ri % 2 == 0)
        ws2.cell(row=ri, column=2).number_format = '₹#,##0.00'
        total_sub += s['amount']

    # Total row
    tr = len(data['subscriptions']) + 2
    ws2.cell(row=tr, column=1, value="TOTAL")
    ws2.cell(row=tr, column=2, value=total_sub)
    for col in range(1, 4):
        cell = ws2.cell(row=tr, column=col)
        cell.fill   = header_fill(C_LIGHT)
        cell.font   = Font(name='Calibri', bold=True, size=9, color=C_PRIMARY)
        cell.border = thin_border()
    ws2.cell(row=tr, column=2).number_format = '₹#,##0.00'

    for col, w in zip([1,2,3], [24, 18, 14]):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ─────────────────────────────────────────────────────────────────
    # Sheet 5 – Investments
    # ─────────────────────────────────────────────────────────────────
    wi = wb.create_sheet("Investments")
    wi.sheet_view.showGridLines = False
    iheaders = ["Name", "Type", "Invested (₹)", "Current Value (₹)", "Gain / Loss (₹)", "Date"]
    for ci, h in enumerate(iheaders, 1):
        wi.cell(row=1, column=ci, value=h)
    style_header_row(wi, 1, len(iheaders))

    for ri, inv in enumerate(data['investments'], 2):
        gain = inv.get('gain', 0)
        row  = [
            inv['name'], inv['type'],
            inv.get('amount', 0),
            inv.get('current_val', 0),
            gain,
            inv.get('invest_date', ''),
        ]
        for ci, val in enumerate(row, 1):
            wi.cell(row=ri, column=ci, value=val)
        style_data_row(wi, ri, len(iheaders), ri % 2 == 0)
        for col in [3, 4, 5]:
            wi.cell(row=ri, column=col).number_format = '₹#,##0.00'
        gain_cell = wi.cell(row=ri, column=5)
        gain_cell.font = normal_font(bold=True, color=C_SUCCESS if gain >= 0 else C_DANGER)

    for col, w in zip([1,2,3,4,5,6], [20, 16, 16, 18, 16, 12]):
        wi.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
